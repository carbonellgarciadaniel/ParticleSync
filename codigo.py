import os
import math
import csv
import openpyxl
import sys
import cv2
import numpy as np
import logging
from PyQt5.QtCore import Qt, QSize
from PyQt5.QtGui import QImage, QPixmap, QIcon
from PyQt5.QtWidgets import (
    QApplication,
    QDialog,
    QFileDialog,
    QInputDialog,
    QLabel,
    QListWidget,
    QMainWindow,
    QMenuBar,
    QMenu,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QTreeWidget,
    QTreeWidgetItem,
    QVBoxLayout,
    QWidget,
    QHBoxLayout,
    QAction,
    QSizePolicy,
    QSpacerItem,
    QSizePolicy,
    QHeaderView,
    QGridLayout
)
from PIL import Image # Agregado import de Pillow

class ImagenZoom:
    def __init__(self, imagen):
        self.imagen = imagen
        self.imagen_original = imagen.copy()
        self.zoom_factor = 1.0
        self.x_offset = 0
        self.y_offset = 0
        self.arrastrando = False
        self.x_anterior = 0
        self.y_anterior = 0

    def mostrar_imagen(self):
        M = np.float32([
            [self.zoom_factor, 0, self.x_offset],
            [0, self.zoom_factor, self.y_offset]
        ])
        alto, ancho = self.imagen.shape[:2]
        imagen_transformada = cv2.warpAffine(self.imagen_original, M, (ancho, alto))
        cv2.imshow("Imagen Procesada", imagen_transformada)

    def mouse_callback(self, event, x, y, flags, param):
        if event == cv2.EVENT_MOUSEWHEEL:
            if flags > 0:
                self.zoom_factor *= 1.1
            else:
                self.zoom_factor /= 1.1
            self.zoom_factor = max(1.0, self.zoom_factor)
            self.mostrar_imagen()
        elif event == cv2.EVENT_LBUTTONDOWN:
            self.arrastrando = True
            self.x_anterior = x
            self.y_anterior = y
        elif event == cv2.EVENT_MOUSEMOVE and self.arrastrando:
            delta_x = x - self.x_anterior
            delta_y = y - self.y_anterior
            self.x_offset += delta_x
            self.y_offset += delta_y
            self.x_anterior = x
            self.y_anterior = y
            self.mostrar_imagen()
        elif event == cv2.EVENT_LBUTTONUP:
            self.arrastrando = False


class ParticleCounterApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("ParticleSync")
        self.setGeometry(300, 300, 1000, 950)  # Ajustado el tamaño de la ventana
        self.setStyleSheet("""
            QMainWindow {
                background-color: #e8f0fe;
                font-family: Arial;
                font-size: 12px;
            }
            QPushButton {
                background-color: #3498db;  /* Color de fondo */
                color: white;               /* Color del texto */
                border: 1px solid #cccccc;  /* Borde */
                border-radius: 5px;         /* Bordes redondeados */
                font-weight: bold;
                padding: 5px 10px;          /* Espaciado interno reducido */
                font-size: 12px;            /* Tamaño de fuente más pequeño */
            }
            QPushButton:hover {
                background-color: #2980b9;  /* Color de fondo al pasar el ratón */
            }
            QListWidget, QTableWidget, QLabel#canvas {
                background-color: white;
                border: 1px solid #cccccc;
            }
            QProgressBar {
                border: 1px solid #cccccc;
                border-radius: 5px;
                text-align: center;
            }

            QProgressBar::chunk {
                background-color: #228B22;
                width: 20px;
            }
        """)
        
        self.rutas_imagenes = []
        self.imagen_procesada = None
        self.um_per_px = 0.5
        self.datos_particulas = []
        self.datos_fibras = []
        self.rango_longitud = (0, float('inf'))

        self.initUI()
        self.initIcon()
        self.cargar_logo()
        
    def cargar_logo(self):
        # Usar una ruta relativa
        logo_ruta = "iconos/Logo_UCA.png"
        if os.path.exists(logo_ruta):
            logo_pixmap = QPixmap(logo_ruta)
            self.lbl_logo = QLabel()  # Inicializar QLabel aquí
            self.lbl_logo.setPixmap(logo_pixmap.scaled(715, 310, Qt.KeepAspectRatio))
            self.lbl_logo.setAlignment(Qt.AlignCenter)
            self.layout_lista.addWidget(self.lbl_logo)
        else:
            logging.error(f"No se encontró el logo en la ruta: {logo_ruta}")
            # Puedes usar un logo predeterminado o mostrar un mensaje de error aquí
            
    def initIcon(self):
        # Usar una ruta relativa
        icono_ruta = "iconos/icono.ico"
        if os.path.exists(icono_ruta): # Verificar si el archivo existe
            icon = QIcon(icono_ruta)
            self.setWindowIcon(icon)
        else:
            logging.error(f"No se encontró el icono en la ruta: {icono_ruta}")
            # Puedes usar un icono predeterminado o mostrar un mensaje de error aquí
        
    def initUI(self):
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QHBoxLayout(self.central_widget)

        # Frame para la lista de imágenes
        self.frame_lista = QWidget()
        self.frame_lista.setFixedWidth(250)  # Reducido el ancho
        self.layout_lista = QVBoxLayout(self.frame_lista)

        self.btn_cargar = QPushButton("Cargar Imágenes")
        self.btn_cargar.clicked.connect(self.cargar_imagenes)
        self.layout_lista.addWidget(self.btn_cargar)

        self.lista_imagenes = QListWidget()
        self.lista_imagenes.itemSelectionChanged.connect(self.mostrar_previsualizacion)
        self.layout_lista.addWidget(self.lista_imagenes)

        self.btn_eliminar = QPushButton("Eliminar Imagen Seleccionada")
        self.btn_eliminar.clicked.connect(self.eliminar_imagen)
        self.layout_lista.addWidget(self.btn_eliminar)

        self.btn_procesar_todas = QPushButton("Procesar Todas las Imágenes")
        self.btn_procesar_todas.clicked.connect(self.procesar_todas_imagenes)
        self.layout_lista.addWidget(self.btn_procesar_todas)

        # Botón para eliminar parámetros predeterminados
        self.btn_eliminar_parametros = QPushButton("Eliminar P. Predeterminados")
        self.btn_eliminar_parametros.clicked.connect(self.eliminar_parametros_predeterminados)
        self.btn_eliminar_parametros.setEnabled(False)  # Inicialmente deshabilitado
        self.layout_lista.addWidget(self.btn_eliminar_parametros)

        # Añadir un espacio flexible
        self.layout_lista.addItem(QSpacerItem(5, 10, QSizePolicy.Minimum, QSizePolicy.Expanding))

        self.layout.addWidget(self.frame_lista)

        # Frame para la previsualización y procesamiento
        self.frame_preview = QWidget()
        self.layout_preview = QGridLayout(self.frame_preview)  # Usar QGridLayout

        # QLabel con estilo y tamaño fijo
        self.lbl_preview = QLabel("Previsualización")
        self.lbl_preview.setAlignment(Qt.AlignCenter)
        self.lbl_preview.setFixedSize(750, 40)  # Ancho: 200px, Alto: 50px
        self.lbl_preview.setStyleSheet("""
            background-color: #3498db;  /* Fondo azul oscuro */
            color: #FFF;               /* Texto en blanco */
            font-size: 16px;           /* Tamaño de fuente más pequeño */
            font-family: 'Arial';      /* Fuente Arial */
            font-weight: bold;  
            border: 2px solid #ADD8E6;    /* Borde gris */
            border-radius: 10px;       /* Bordes redondeados */
            padding: 10px;             /* Espaciado interno */
        """)

        # Centrar el QLabel en el QGridLayout
        self.layout_preview.addWidget(self.lbl_preview, 0, 0, alignment=Qt.AlignCenter)
        
        # Canvas para mostrar la imagen
        self.canvas = QLabel()
        self.canvas.setFixedSize(1500, 700)  # Reducido el tamaño del canvas
        self.canvas.setStyleSheet("background-color: #e8f0fe;")  # Color de fondo solo para el canvas
        self.layout_preview.addWidget(self.canvas)

        self.progress = QProgressBar()
        self.layout_preview.addWidget(self.progress)

        self.btn_procesar = QPushButton("Procesar Imagen Seleccionada")
        self.btn_procesar.clicked.connect(self.procesar_imagen)
        self.btn_procesar.setEnabled(False)
        self.layout_preview.addWidget(self.btn_procesar)

        self.btn_guardar = QPushButton("Guardar Imagen Procesada")
        self.btn_guardar.clicked.connect(self.guardar_imagen)
        self.btn_guardar.setEnabled(False)
        self.layout_preview.addWidget(self.btn_guardar)

        # Estilo personalizado para el botón de ayuda
        estilo_ayuda = """
            QPushButton {
                background-color: #FFC107; /* Amarillo suave o naranja */
                color: black; /* Texto en negro para mejor contraste */
                border: 1px solid #FF9800; /* Borde ligeramente más oscuro */
                padding: 5px 10px;
                font-weight: bold;
                border-radius: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #FFD54F; /* Un tono más claro al pasar el ratón */
            }
        """
        self.btn_ayuda = QPushButton("Ayuda")
        self.btn_ayuda.clicked.connect(self.mostrar_ayuda)
        self.btn_ayuda.setStyleSheet(estilo_ayuda)
        self.layout_preview.addWidget(self.btn_ayuda)
        self.btn_ayuda.setFixedWidth(100)  # Establecer ancho fijo
        self.btn_ayuda.setFixedHeight(30)  # Establecer alto fijo
        self.layout.addWidget(self.frame_preview)
        # Centrar el botón "Ayuda" en la fila 5, columna 0
        self.layout_preview.addWidget(self.btn_ayuda, 5, 0, alignment=Qt.AlignCenter)
        self.initMenu()

    def cargar_imagenes(self):
        rutas, _ = QFileDialog.getOpenFileNames(self, "Seleccionar imágenes", "", "Imágenes (*.png *.tiff)")
        if rutas:
            self.rutas_imagenes.extend(rutas)
            self.lista_imagenes.clear()
            for ruta in self.rutas_imagenes:
                self.lista_imagenes.addItem(os.path.basename(ruta))
            self.btn_procesar.setEnabled(True)

    def mostrar_previsualizacion(self):
        seleccion = self.lista_imagenes.currentRow()
        if seleccion >= 0:
            ruta = self.rutas_imagenes[seleccion]
            try:
                img_cv = cv2.imread(ruta)
                if img_cv is None:
                    raise Exception("No se pudo cargar la imagen con OpenCV.")

                canvas_width = self.canvas.width()
                canvas_height = self.canvas.height()

                img_height, img_width, _ = img_cv.shape
                aspect_ratio = img_width / img_height

                if aspect_ratio > canvas_width / canvas_height:
                    new_width = canvas_width
                    new_height = int(canvas_width / aspect_ratio)
                else:
                    new_height = canvas_height
                    new_width = int(canvas_height * aspect_ratio)

                img_resized = cv2.resize(img_cv, (new_width, new_height))
                img_rgb = cv2.cvtColor(img_resized, cv2.COLOR_BGR2RGB)

                height, width, channel = img_rgb.shape
                bytes_per_line = 3 * width
                q_img = QImage(img_rgb.data, width, height, bytes_per_line, QImage.Format_RGB888)
                pixmap = QPixmap.fromImage(q_img)

                self.canvas.setPixmap(pixmap)
                self.canvas.setAlignment(Qt.AlignCenter)

            except Exception as e:
                print(f"Error al previsualizar la imagen: {e}")
                self.canvas.setText("Error al cargar la imagen.")
                self.canvas.setAlignment(Qt.AlignCenter)

    def initMenu(self):
        menubar = self.menuBar()

        # Aplicar estilo al menú
        menubar.setStyleSheet("""
            QMenuBar {
                background-color: #a9c2e9;  /* Color de fondo del menú */
                color: #000000;            /* Color del texto */
                font-size: 14px;            /* Tamaño de la fuente */
                font-family: 'Arial';       /* Fuente */
            }
            QMenuBar::item {
                background-color: transparent;  /* Fondo transparente para los ítems */
                padding: 5px 10px;              /* Espaciado interno */
            }
            QMenuBar::item:selected {
                background-color: #c6d9f1;  /* Color de fondo cuando se selecciona un ítem */
            }
            QMenu {
                background-color: #b8cce4;  /* Color de fondo del menú desplegable */
                color: #000000;             /* Color del texto */
                border: 1px solid #c6d9f1;  /* Borde del menú desplegable */
            }
            QMenu::item {
                padding: 5px 20px;          /* Espaciado interno de los ítems del menú */
            }
            QMenu::item:selected {
                background-color: #c6d9f1;  /* Color de fondo cuando se selecciona un ítem */
            }
        """)

        # Menú Archivo
        archivo_menu = menubar.addMenu("Archivo")
        archivo_menu.addAction("Cargar Imágenes", self.cargar_imagenes)
        archivo_menu.addSeparator()
        archivo_menu.addAction("Salir", self.close)

        # Menú Herramientas
        herramientas_menu = menubar.addMenu("Herramientas")
        herramientas_menu.addAction("Medir Distancia", self.medir_distancia)
        herramientas_menu.addAction("Conteo Manual", self.conteo_manual)
        herramientas_menu.addAction("Medir Area e Intensidad", self.medir_area_intensidad)
        herramientas_menu.addSeparator()
        herramientas_menu.addAction("Establecer Parámetros Predeterminados", self.establecer_parametros_predeterminados)  # Nueva opción

        # Menú Ayuda
        ayuda_menu = menubar.addMenu("Ayuda")
        ayuda_menu.addAction("Ayuda", self.mostrar_ayuda)

        # Menú Acerca de
        acerca_de_menu = menubar.addMenu("Acerca de")
        acerca_de_action = QAction("Acerca de...", self)  # Crear la acción
        acerca_de_action.triggered.connect(self.mostrar_acerca_de)  # Conectar la acción a la función
        acerca_de_menu.addAction(acerca_de_action)

    def eliminar_parametros_predeterminados(self):
        # Verificar si los parámetros predeterminados existen
        if hasattr(self, 'escala_predeterminada') or hasattr(self, 'rango_longitud_predeterminado'):
            # Eliminar los atributos
            if hasattr(self, 'escala_predeterminada'):
                delattr(self, 'escala_predeterminada')
            if hasattr(self, 'rango_longitud_predeterminado'):
                delattr(self, 'rango_longitud_predeterminado')

            # Deshabilitar el botón
            self.btn_eliminar_parametros.setEnabled(False)

            # Mostrar un mensaje de confirmación
            QMessageBox.information(self, "Parámetros Eliminados", "Los parámetros predeterminados han sido eliminados.")
        else:
            # Si no hay parámetros predeterminados, mostrar un mensaje
            QMessageBox.information(self, "Sin Parámetros", "No hay parámetros predeterminados para eliminar.")

    def pedir_escala_y_rango(self):
        # Pedir la escala en µm/px
        um_per_px, ok = QInputDialog.getDouble(self, "Escala", "Ingresa la escala en µm/px:", decimals=2)
        if not ok:
            return

        um_per_px_str = str(um_per_px).replace(",", ".")  # Reemplazar comas por puntos
        try:
            um_per_px = float(um_per_px_str)
        except ValueError:
            QMessageBox.critical(self, "Error", "Escala inválida. Por favor, ingresa un número válido.")
            return

        if um_per_px <= 0:
            QMessageBox.critical(self, "Error", "La escala debe ser mayor que cero.")
            return

        self.um_per_px = um_per_px

        # Pedir el rango de longitud en µm
        longitud_min, ok = QInputDialog.getDouble(self, "Rango de Longitud", "Ingresa la longitud mínima en µm:", min=0)
        if not ok:
            return
        longitud_max, ok = QInputDialog.getDouble(self, "Rango de Longitud", "Ingresa la longitud máxima en µm:", min=longitud_min)
        if not ok:
            return

        self.rango_longitud = (longitud_min, longitud_max)

    def establecer_parametros_predeterminados(self):
        # Pedir la escala en µm/px
        um_per_px, ok = QInputDialog.getDouble(self, "Escala Predeterminada", "Ingresa la escala en µm/px:", value=self.um_per_px, decimals=2)
        if not ok:
            return

        um_per_px_str = str(um_per_px).replace(",", ".")  # Reemplazar comas por puntos
        try:
            um_per_px = float(um_per_px_str)
        except ValueError:
            QMessageBox.critical(self, "Error", "Escala inválida. Por favor, ingresa un número válido.")
            return

        if um_per_px <= 0:
            QMessageBox.critical(self, "Error", "La escala debe ser mayor que cero.")
            return

        # Pedir el rango de longitud en µm
        longitud_min, ok = QInputDialog.getDouble(self, "Rango de Longitud Predeterminado", "Ingresa la longitud mínima en µm:", min=0)
        if not ok:
            return
        longitud_max, ok = QInputDialog.getDouble(self, "Rango de Longitud Predeterminado", "Ingresa la longitud máxima en µm:", min=longitud_min)
        if not ok:
            return

        # Guardar los valores predeterminados
        self.escala_predeterminada = um_per_px
        self.rango_longitud_predeterminado = (longitud_min, longitud_max)

        # Habilitar el botón de eliminar parámetros
        self.btn_eliminar_parametros.setEnabled(True)

        QMessageBox.information(self, "Parámetros Predeterminados", "Los parámetros predeterminados se han establecido correctamente.")
                
    def cargar_imagenes(self): # Función definida correctamente
        rutas, _ = QFileDialog.getOpenFileNames(self, "Seleccionar imágenes", "", "Imágenes (*.png *.tiff)")
        if rutas:
            self.rutas_imagenes.extend(rutas)
            self.lista_imagenes.clear()
            for ruta in self.rutas_imagenes:
                self.lista_imagenes.addItem(os.path.basename(ruta))
            self.btn_procesar.setEnabled(True)

    def mostrar_previsualizacion(self):
        seleccion = self.lista_imagenes.currentRow()
        if seleccion >= 0 and seleccion < len(self.rutas_imagenes):  # Verificar que el índice sea válido
            ruta = self.rutas_imagenes[seleccion]
            try:
                img_cv = cv2.imread(ruta)
                if img_cv is None:
                    raise Exception("No se pudo cargar la imagen con OpenCV.")

                canvas_width = self.canvas.width()
                canvas_height = self.canvas.height()

                img_height, img_width, _ = img_cv.shape
                aspect_ratio = img_width / img_height

                if aspect_ratio > canvas_width / canvas_height:
                    new_width = canvas_width
                    new_height = int(canvas_width / aspect_ratio)
                else:
                    new_height = canvas_height
                    new_width = int(canvas_height * aspect_ratio)

                img_resized = cv2.resize(img_cv, (new_width, new_height))
                img_rgb = cv2.cvtColor(img_resized, cv2.COLOR_BGR2RGB)

                height, width, channel = img_rgb.shape
                bytes_per_line = 3 * width
                q_img = QImage(img_rgb.data, width, height, bytes_per_line, QImage.Format_RGB888)
                pixmap = QPixmap.fromImage(q_img)

                self.canvas.setPixmap(pixmap)
                self.canvas.setAlignment(Qt.AlignCenter)

            except Exception as e:
                print(f"Error al previsualizar la imagen: {e}")
                self.canvas.setText("Error al cargar la imagen.")
                self.canvas.setAlignment(Qt.AlignCenter)
        else:
            # Si no hay una selección válida, limpiar el canvas
            self.canvas.clear()
            self.canvas.setText("No hay imagen seleccionada.")
            self.canvas.setAlignment(Qt.AlignCenter)
            
    def eliminar_imagen(self):
        seleccion = self.lista_imagenes.currentRow()
        if seleccion >= 0 and seleccion < len(self.rutas_imagenes):  # Verificar que el índice sea válido
            try:
                logging.debug(f"Eliminando imagen en índice: {seleccion}")
                logging.debug(f"Rutas antes de eliminar: {self.rutas_imagenes}")
                logging.debug(f"Lista antes de eliminar: {[self.lista_imagenes.item(i).text() for i in range(self.lista_imagenes.count())]}")

                self.rutas_imagenes.pop(seleccion)
                self.lista_imagenes.takeItem(seleccion)

                logging.debug(f"Rutas después de eliminar: {self.rutas_imagenes}")
                logging.debug(f"Lista después de eliminar: {[self.lista_imagenes.item(i).text() for i in range(self.lista_imagenes.count())]}")

                if not self.rutas_imagenes:
                    self.btn_procesar.setEnabled(False)
                    self.btn_guardar.setEnabled(False)
            except Exception as e:
                logging.error(f"Error al eliminar la imagen: {e}")
                QMessageBox.critical(self, "Error", f"Ocurrió un error al eliminar la imagen: {e}")
        else:
            QMessageBox.warning(self, "Advertencia", "No hay imagen seleccionada para eliminar.")

    def procesar_imagen(self):
        try:
            # Obtener la imagen seleccionada
            seleccion = self.lista_imagenes.currentRow()
            if seleccion < 0:
                QMessageBox.critical(self, "Error", "Por favor, selecciona una imagen de la lista.")
                return

            self.ruta_imagen = self.rutas_imagenes[seleccion]
            self.progress.setValue(0)
            
            # Preguntar si se desea usar los valores predeterminados
            if hasattr(self, 'escala_predeterminada') and hasattr(self, 'rango_longitud_predeterminado'):
                usar_predeterminados = QMessageBox.question(
                    self,
                    "Usar Parámetros Predeterminados",
                    "¿Deseas usar los parámetros predeterminados?\n"
                    f"Escala: {self.escala_predeterminada} µm/px\n"
                    f"Rango de longitud: {self.rango_longitud_predeterminado[0]} - {self.rango_longitud_predeterminado[1]} µm",
                    QMessageBox.Yes | QMessageBox.No,
                )
                if usar_predeterminados == QMessageBox.Yes:
                    self.um_per_px = self.escala_predeterminada
                    self.rango_longitud = self.rango_longitud_predeterminado
                else:
                    # Pedir la escala y el rango de longitud manualmente
                    self.pedir_escala_y_rango()
            else:
                # Si no hay valores predeterminados, pedir la escala y el rango de longitud manualmente
                self.pedir_escala_y_rango()

            # Pedir el tipo de preprocesamiento
            opcion, ok = QInputDialog.getInt(self, "Tipo de Preprocesamiento", 
                                             "Elige el tipo de preprocesamiento:\n1. Partículas blancas sobre fondo negro\n2. Partículas negras sobre fondo blanco",
                                             min=1, max=2)
            if not ok:
                return

            # Cargar la imagen
            imagen = cv2.imread(self.ruta_imagen, cv2.IMREAD_GRAYSCALE)
            if imagen is None:
                QMessageBox.critical(self, "Error", "No se pudo cargar la imagen.")
                return

            # --- PREPROCESAMIENTO DE LA IMAGEN ---
            if opcion == 1:
                # Procesamiento para partículas blancas sobre fondo negro
                imagen_eq = cv2.equalizeHist(imagen)  # Ecualización del histograma para mejorar el contraste
                imagen_blur = cv2.bilateralFilter(imagen_eq, 5, 50, 50)  # Filtro bilateral para reducir ruido preservando bordes
                _, umbralizada = cv2.threshold(imagen, 127, 255, cv2.THRESH_BINARY + cv2.THRESH_TRIANGLE)
            elif opcion == 2:
                # Procesamiento para partículas negras sobre fondo blanco
                _, umbralizada = cv2.threshold(imagen, 45, 255, cv2.THRESH_BINARY_INV)  # Umbralización adaptativa

            # Operación de cierre para rellenar huecos
            kernel = np.ones((3, 3), np.uint8)
            umbralizada = cv2.morphologyEx(umbralizada, cv2.MORPH_CLOSE, kernel)  # Cierre para rellenar huecos
            # --- FIN PREPROCESAMIENTO ---

            self.progress.setValue(40)

            # --- DETECCIÓN DE CONTORNOS ---
            contornos, _ = cv2.findContours(umbralizada, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)  # Usar la imagen umbralizada
            imagen_color = cv2.cvtColor(imagen, cv2.COLOR_GRAY2BGR)
            contornos_aprox = []  # Lista para almacenar contornos simplificados

            # Definir el rango de longitud para las fibras (en µm)
            rango_fibras = (350, 1500)  # Fibras entre 350 µm y 1500 µm
            self.datos_particulas = []
            self.datos_fibras = []

            for i, contorno in enumerate(contornos, start=1):
                # Calcular área y perímetro
                area_px = cv2.contourArea(contorno)
                perimetro_px = cv2.arcLength(contorno, True)
                area_um2 = area_px * (self.um_per_px ** 2)
                perimetro_um = perimetro_px * self.um_per_px

                # --- SIMPLIFICACIÓN DE CONTORNOS ---
                epsilon = 0.00001 * cv2.arcLength(contorno, True)  # Ajusta el valor de epsilon
                contorno_aprox = cv2.approxPolyDP(contorno, epsilon, True)
                contornos_aprox.append(contorno_aprox)

                if area_um2 > 1:
                    # Calcular el diámetro máximo utilizando el círculo mínimo que encierra el contorno
                    (_, _), radius = cv2.minEnclosingCircle(contorno)
                    diametro_max_um = 2 * radius * self.um_per_px

                    # Calcular el diámetro mínimo utilizando el rectángulo de área mínima
                    rect = cv2.minAreaRect(contorno)
                    ancho, alto = rect[1]
                    diametro_min_um = min(ancho, alto) * self.um_per_px

                    # Calcular relación de aspecto y circularidad
                    x, y, w, h = cv2.boundingRect(contorno)
                    relacion_aspecto = diametro_min_um / diametro_max_um if diametro_max_um != 0 else 0
                    circularidad = (4 * np.pi * area_px) / (cv2.arcLength(contorno, True) ** 2) if cv2.arcLength(contorno, True) != 0 else 0

                    # Filtrar por rango de longitud
                    if relacion_aspecto < 0.3 or circularidad < 0.12:  # Fibras
                        if rango_fibras[0] <= diametro_max_um <= rango_fibras[1]:
                            self.datos_fibras.append((i, diametro_max_um, diametro_min_um, area_um2, perimetro_um, relacion_aspecto, circularidad))
                            cv2.drawContours(imagen_color, [contorno_aprox], -1, (0, 0, 255), 2)  # Rojo para fibras
                            cv2.putText(imagen_color, str(i), (x + w // 2, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 255), 2)
                    else:  # Partículas
                        if self.rango_longitud[0] <= diametro_max_um <= self.rango_longitud[1]:
                            # Calcular la intensidad media de la partícula
                            mascara = np.zeros_like(imagen)
                            cv2.drawContours(mascara, [contorno], -1, 255, -1)
                            intensidad_media = cv2.mean(imagen, mask=mascara)[0]

                            # Aplicar el umbral de intensidad solo para partículas blancas sobre fondo negro
                            if opcion == 1 and intensidad_media < 63:
                                continue  # Ignorar partículas que no cumplen con el umbral

                            # Agregar partícula a los datos
                            self.datos_particulas.append((i, diametro_max_um, diametro_min_um, area_um2, perimetro_um, relacion_aspecto, circularidad))
                            cv2.drawContours(imagen_color, [contorno_aprox], -1, (0, 255, 0), 2)  # Verde para partículas
                            cv2.putText(imagen_color, str(i), (x + w // 2, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 255), 2)

            # Calcular intensidades para partículas y fibras
            intensidades_particulas = self.calcular_intensidades(imagen, [contorno for contorno in contornos_aprox if self.rango_longitud[0] <= cv2.minEnclosingCircle(contorno)[1] * 2 * self.um_per_px <= self.rango_longitud[1]])
            intensidades_fibras = self.calcular_intensidades(imagen, [contorno for contorno in contornos_aprox if 350 <= cv2.minEnclosingCircle(contorno)[1] * 2 * self.um_per_px <= 1500])

            # Agregar intensidades a los datos de partículas y fibras
            for j, dato in enumerate(self.datos_particulas):
                media, maxima, minima = intensidades_particulas[j]
                self.datos_particulas[j] = dato + (media, maxima, minima)

            for j, dato in enumerate(self.datos_fibras):
                media, maxima, minima = intensidades_fibras[j]
                self.datos_fibras[j] = dato + (media, maxima, minima)

            self.progress.setValue(80)

            # Mostrar la imagen procesada
            self.imagen_procesada = imagen_color

            # Variables para el zoom, desplazamiento y estado de arrastre
            zoom_factor = 1.0
            x_offset, y_offset = 0, 0
            arrastrando = False
            x_inicio, y_inicio = 0, 0

            def actualizar_imagen():
                nonlocal imagen_color, zoom_factor, x_offset, y_offset

                # Crear una matriz de transformación para el zoom y el desplazamiento
                M = np.float32([
                    [zoom_factor, 0, x_offset],
                    [0, zoom_factor, y_offset]
                ])

                # Aplicar la transformación a la imagen original
                alto, ancho = imagen_color.shape[:2]
                imagen_transformada = cv2.warpAffine(imagen_color, M, (ancho, alto))

                # Redimensionar la imagen para que se ajuste a la ventana
                alto_ventana, ancho_ventana = 800, 1200  # Tamaño de la ventana
                imagen_redimensionada = cv2.resize(imagen_transformada, (ancho_ventana, alto_ventana))

                # Mostrar la imagen actualizada
                cv2.imshow("Imagen Procesada", imagen_redimensionada)

            def clic_derecho(event, x, y, flags, param):
                nonlocal zoom_factor, x_offset, y_offset, arrastrando, x_inicio, y_inicio

                # Ajustar las coordenadas del mouse al tamaño de la imagen original
                x_original = int(x * (imagen_color.shape[1] / 1200))  # 1200 es el ancho de la ventana
                y_original = int(y * (imagen_color.shape[0] / 800))   # 800 es el alto de la ventana

                if event == cv2.EVENT_MOUSEWHEEL:
                    # Hacer zoom con la rueda del ratón
                    if flags > 0:  # Rueda hacia arriba (zoom in)
                        nuevo_zoom = zoom_factor * 1.1
                    else:  # Rueda hacia abajo (zoom out)
                        nuevo_zoom = zoom_factor / 1.1
                    nuevo_zoom = max(1.0, nuevo_zoom)  # Evitar zoom menor a 1

                    # Ajustar los offsets para que el zoom se centre en el cursor
                    x_raton, y_raton = x_original, y_original
                    x_offset = x_raton - (x_raton - x_offset) * (nuevo_zoom / zoom_factor)
                    y_offset = y_raton - (y_raton - y_offset) * (nuevo_zoom / zoom_factor)
                    zoom_factor = nuevo_zoom

                    actualizar_imagen()

                elif event == cv2.EVENT_LBUTTONDOWN:
                    # Iniciar arrastre
                    arrastrando = True
                    x_inicio, y_inicio = x_original, y_original

                elif event == cv2.EVENT_MOUSEMOVE:
                    if arrastrando:
                        # Calcular el desplazamiento
                        delta_x = x_original - x_inicio
                        delta_y = y_original - y_inicio
                        x_offset += delta_x
                        y_offset += delta_y
                        x_inicio, y_inicio = x_original, y_original
                        actualizar_imagen()

                elif event == cv2.EVENT_LBUTTONUP:
                    # Detener arrastre
                    arrastrando = False

            cv2.namedWindow("Imagen Procesada", cv2.WINDOW_NORMAL)
            cv2.resizeWindow("Imagen Procesada", 1200, 800)  # Tamaño de la ventana
            cv2.setMouseCallback("Imagen Procesada", clic_derecho)  # Enlazar evento del ratón
            actualizar_imagen()  # Mostrar la imagen inicial
            cv2.waitKey(1)  # Esperar hasta que se cierre la ventana

            # Mostrar la tabla y el resumen
            self.mostrar_tabla()
            self.btn_guardar.setEnabled(True)
            self.mostrar_resumen()
            self.progress.setValue(100)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Ocurrió un error: {e}")

    def mostrar_imagen_procesada(self, imagen_color):
        # Variables para el zoom, desplazamiento y estado de arrastre
        zoom_factor = 1.0
        x_offset, y_offset = 0, 0
        arrastrando = False
        x_inicio, y_inicio = 0, 0

        def actualizar_imagen():
            nonlocal imagen_color, zoom_factor, x_offset, y_offset
            
    def calcular_intensidades(self, imagen_gris, contornos):
        resultados = []
        for contorno in contornos:
            mascara = np.zeros_like(imagen_gris)
            cv2.drawContours(mascara, [contorno], 0, 255, cv2.FILLED)
            pixeles = imagen_gris[mascara == 255]
            if pixeles.size > 0:
                intensidad_media = np.mean(pixeles)
                intensidad_maxima = np.max(pixeles)
                intensidad_minima = np.min(pixeles)
                resultados.append((intensidad_media, intensidad_maxima, intensidad_minima))
            else:
                resultados.append((0, 0, 0))
        return resultados

    def mostrar_tabla(self):
        ventana_tabla = QDialog(self)
        ventana_tabla.setWindowTitle("Resultados y Resumen")
        ventana_tabla.resize(1200, 800)
        layout = QVBoxLayout(ventana_tabla)

        notebook = QTabWidget()
        layout.addWidget(notebook)

        # Función para guardar datos en Excel
        def guardar_excel(datos, nombre_archivo):
            try:
                libro_excel = openpyxl.Workbook()
                hoja_excel = libro_excel.active
                # Escribir encabezado
                encabezado = ["ID", "Diámetro (µm)", "Diámetro Mín (µm)","Área (µm²)", "Perímetro (µm)", "Relación de Aspecto", "Circularidad"]
                hoja_excel.append(encabezado)
                # Escribir datos
                for fila in datos:
                    hoja_excel.append(fila)
                libro_excel.save(nombre_archivo)
                QMessageBox.information(ventana_tabla, "Éxito", f"Datos guardados en {nombre_archivo}")
            except Exception as e:
                QMessageBox.critical(ventana_tabla, "Error", f"No se pudieron guardar los datos: {e}")

        # Pestaña de partículas
        frame_particulas = QWidget()
        layout_particulas = QVBoxLayout(frame_particulas)
        tabla_particulas = QTableWidget()
        tabla_particulas.setColumnCount(7)
        tabla_particulas.setHorizontalHeaderLabels(["ID", "Diámetro (µm)", "Diámetro Mín (µm)","Área (µm²)", "Perímetro (µm)", "Relación de Aspecto", "Circularidad"])
        # Ajustar el ancho de las columnas
        tabla_particulas.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        # Ajustar la altura de las filas
        tabla_particulas.verticalHeader().setDefaultSectionSize(30)
        for dato in self.datos_particulas:
            tabla_particulas.insertRow(tabla_particulas.rowCount())
            for i, valor in enumerate(dato):
                tabla_particulas.setItem(tabla_particulas.rowCount() - 1, i, QTableWidgetItem(str(valor)))
        layout_particulas.addWidget(tabla_particulas)

        # Botón para guardar partículas en Excel
        btn_guardar_particulas = QPushButton("Guardar Excel")
        btn_guardar_particulas.clicked.connect(lambda: guardar_excel(self.datos_particulas, QFileDialog.getSaveFileName(ventana_tabla, "Guardar Partículas en Excel", "", "Archivos Excel (*.xlsx)")[0]))
        layout_particulas.addWidget(btn_guardar_particulas)

        notebook.addTab(frame_particulas, "Partículas")

        # Pestaña de fibras
        frame_fibras = QWidget()
        layout_fibras = QVBoxLayout(frame_fibras)
        tabla_fibras = QTableWidget()
        tabla_fibras.setColumnCount(7)
        tabla_fibras.setHorizontalHeaderLabels(["ID", "Diámetro (µm)", "Diámetro Mín (µm)","Área (µm²)", "Perímetro (µm)", "Relación de Aspecto", "Circularidad"])
        # Ajustar el ancho de las columnas
        tabla_fibras.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        # Ajustar la altura de las filas
        tabla_fibras.verticalHeader().setDefaultSectionSize(30)
        for dato in self.datos_fibras:
            tabla_fibras.insertRow(tabla_fibras.rowCount())
            for i, valor in enumerate(dato):
                tabla_fibras.setItem(tabla_fibras.rowCount() - 1, i, QTableWidgetItem(str(valor)))
        layout_fibras.addWidget(tabla_fibras)

        # Botón para guardar fibras en Excel
        btn_guardar_fibras = QPushButton("Guardar Excel")
        btn_guardar_fibras.clicked.connect(lambda: guardar_excel(self.datos_fibras, QFileDialog.getSaveFileName(ventana_tabla, "Guardar Fibras en Excel", "", "Archivos Excel (*.xlsx)")[0]))
        layout_fibras.addWidget(btn_guardar_fibras)

        notebook.addTab(frame_fibras, "Fibras")

        # Pestaña de resumen
        frame_resumen = QWidget()
        layout_resumen = QVBoxLayout(frame_resumen)
        lbl_particulas = QLabel(f"Partículas detectadas: {len(self.datos_particulas)}")
        lbl_fibras = QLabel(f"Fibras detectadas: {len(self.datos_fibras)}")
        layout_resumen.addWidget(lbl_particulas)
        layout_resumen.addWidget(lbl_fibras)
        notebook.addTab(frame_resumen, "Resumen")

        ventana_tabla.exec_()
    
    def mostrar_resumen(self):
        ventana_resumen = QDialog(self)
        ventana_resumen.setWindowTitle("Resumen de Conteo")
        layout = QVBoxLayout(ventana_resumen)

        lbl_particulas = QLabel(f"Partículas detectadas: {len(self.datos_particulas)}")
        lbl_fibras = QLabel(f"Fibras detectadas: {len(self.datos_fibras)}")
        layout.addWidget(lbl_particulas)
        layout.addWidget(lbl_fibras)

        ventana_resumen.exec_()

    def guardar_imagen(self):
        if self.imagen_procesada is None:
            QMessageBox.critical(self, "Error", "No hay imagen procesada para guardar.")
            return

        ruta_guardado, _ = QFileDialog.getSaveFileName(self, "Guardar imagen", "", "PNG (*.png);;TIFF (*.tiff)")
        if ruta_guardado:
            cv2.imwrite(ruta_guardado, self.imagen_procesada)
            QMessageBox.information(self, "Éxito", "Imagen guardada correctamente.")

    from PyQt5.QtWidgets import QMessageBox

    def mostrar_ayuda(self):
        ayuda_texto = ayuda_texto = """
<h3>Instrucciones de uso:</h3>
<ol>
<li><b>Cargar Imágenes:</b><br>
- Haz clic en el botón 'Cargar Imágenes' para seleccionar varias imágenes en formatos compatibles (PNG, TIFF, JPG, BMP).<br>
- <i>Nuevo:</i> Ahora puedes cargar un máximo de 20 imágenes a la vez (límite ampliado).<br><br>
<li><b>Seleccionar Imagen:</b><br>
- Una vez cargadas, las imágenes aparecerán en una lista. Selecciona una de la lista para previsualizarla en la ventana principal.<br><br>
<li><b>Procesar Imagen:</b><br>
- Haz clic en 'Procesar Imagen Seleccionada' para iniciar el análisis de la imagen.<br>
- El programa detectará automáticamente partículas o fibras, resaltándolas en la previsualización.<br>
- <i>Nuevo:</i> Ajusta parámetros avanzados como el umbral de sensibilidad y el tamaño mínimo/máximo de partículas para mejorar la precisión.<br><br>
<li><b>Guardar Imagen Procesada:</b><br>
- Después del procesamiento, puedes guardar la imagen con las partículas o fibras detectadas utilizando el botón 'Guardar Imagen'.<br>
- Elige la ubicación y el formato de salida (PNG, TIFF, JPG o BMP).<br><br>
<h3>Recomendaciones para un uso óptimo:</h3>
<ul>
<li>Utiliza imágenes en escala de grises para mejorar la precisión del análisis.</li>
<li>Asegúrate de que las partículas o fibras estén bien definidas y separadas en la imagen.</li>
<li>Evita imágenes con ruido excesivo o fondos no uniformes, ya que pueden afectar la detección.</li>
<li>Si es necesario, utiliza herramientas de preprocesamiento (como ajustes de contraste o filtros) antes de cargar las imágenes.</li>
</ul>
"""

        # Crear un QMessageBox
        msg = QMessageBox(self)
        msg.setWindowTitle("Ayuda")
        msg.setIcon(QMessageBox.Information)
        msg.setTextFormat(1)  # 1 para formato HTML
        msg.setText(ayuda_texto)

        # Mostrar el QMessageBox
        msg.exec_()
    
    def mostrar_acerca_de(self):
        QMessageBox.information(self, "Acerca de...", "ParticleSync Version 1.0 Autor: Daniel Garcia Carbonell")

    def medir_distancia(self):
        try:
            seleccion = self.lista_imagenes.currentRow()
            if seleccion < 0:
                QMessageBox.critical(self, "Error", "Por favor, selecciona una imagen de la lista.")
                return

            ruta = self.rutas_imagenes[seleccion]
            imagen = cv2.imread(ruta, cv2.IMREAD_COLOR)
            if imagen is None:
                QMessageBox.critical(self, "Error", "No se pudo cargar la imagen.")
                return

            # Pedir la escala en µm/px
            um_per_px, ok = QInputDialog.getDouble(self, "Escala", "Ingresa la escala en µm/px:", decimals=2)
            if not ok:
                return

            um_per_px_str = str(um_per_px).replace(",", ".") #Reemplazar comas por puntos.
            try:
                um_per_px = float(um_per_px_str)
            except ValueError:
                QMessageBox.critical(self, "Error", "Escala inválida. Por favor, ingresa un número válido.")
                return

            if um_per_px <= 0:
                QMessageBox.critical(self, "Error", "La escala debe ser mayor que cero.")
                return

            self.um_per_px = um_per_px

            cv2.namedWindow("Medir Distancia", cv2.WINDOW_NORMAL)
            puntos = []
            lineas = []
            textos = []
            imagen_original = imagen.copy()
            zoom_factor = 1.0
            x_offset, y_offset = 0, 0
            arrastrando = False
            x_anterior, y_anterior = 0, 0

            def transformar_coordenadas(x, y):
                return int(x * zoom_factor + x_offset), int(y * zoom_factor + y_offset)

            def invertir_transformacion(x, y):
                return int((x - x_offset) / zoom_factor), int((y - y_offset) / zoom_factor)

            def dibujar_barras(imagen, x1, y1, x2, y2, color, grosor):
                longitud_barra = 10
                cv2.line(imagen, (x1 - longitud_barra, y1), (x1 + longitud_barra, y1), color, grosor)
                cv2.line(imagen, (x1, y1 - longitud_barra), (x1, y1 + longitud_barra), color, grosor)
                cv2.line(imagen, (x2 - longitud_barra, y2), (x2 + longitud_barra, y2), color, grosor)
                cv2.line(imagen, (x2, y2 - longitud_barra), (x2, y2 + longitud_barra), color, grosor)

            def actualizar_imagen():
                nonlocal imagen_original, zoom_factor, x_offset, y_offset, puntos, lineas, textos

                M = np.float32([
                    [zoom_factor, 0, x_offset],
                    [0, zoom_factor, y_offset]
                ])
                alto, ancho = imagen_original.shape[:2]
                imagen_transformada = cv2.warpAffine(imagen_original, M, (ancho, alto))

                for linea in lineas:
                    x1, y1, x2, y2, color, grosor = linea
                    x1_t, y1_t = transformar_coordenadas(x1, y1)
                    x2_t, y2_t = transformar_coordenadas(x2, y2)
                    cv2.line(imagen_transformada, (x1_t, y1_t), (x2_t, y2_t), color, grosor)
                    dibujar_barras(imagen_transformada, x1_t, y1_t, x2_t, y2_t, color, grosor)

                for texto in textos:
                    x, y, texto_distancia, color, fuente, escala = texto
                    x_t, y_t = transformar_coordenadas(x, y)
                    cv2.putText(imagen_transformada, texto_distancia, (x_t, y_t), fuente, escala, color, 2)

                if len(puntos) == 1:
                    x1, y1 = puntos[0]
                    x2, y2 = invertir_transformacion(x_anterior, y_anterior)
                    x1_t, y1_t = transformar_coordenadas(x1, y1)
                    x2_t, y2_t = transformar_coordenadas(x2, y2)
                    cv2.line(imagen_transformada, (x1_t, y1_t), (x2_t, y2_t), (0, 0, 255), 2)
                    dibujar_barras(imagen_transformada, x1_t, y1_t, x2_t, y2_t, (0, 0, 255), 2)

                cv2.imshow("Medir Distancia", imagen_transformada)

            def clic_derecho(event, x, y, flags, param):
                nonlocal puntos, lineas, textos, arrastrando, x_anterior, y_anterior, x_offset, y_offset, zoom_factor

                x_original, y_original = invertir_transformacion(x, y)

                if event == cv2.EVENT_LBUTTONDOWN:
                    if len(puntos) == 0:
                        puntos.append((x_original, y_original))
                    elif len(puntos) == 1:
                        puntos.append((x_original, y_original))
                        x1, y1 = puntos[0]
                        x2, y2 = puntos[1]

                        distancia_px = np.sqrt((x2 - x1)**2 + (y2 - y1)**2)
                        distancia_um = distancia_px * um_per_px

                        lineas.append((x1, y1, x2, y2, (0, 0, 255), 2))

                        texto_distancia = f"{distancia_um:.2f} um"
                        # Calcular el vector perpendicular a la línea
                        dx = x2 - x1
                        dy = y2 - y1
                        longitud = np.sqrt(dx**2 + dy**2)

                        dx /= longitud
                        dy /= longitud

                        # Rotar 90 grados
                        dx_perpendicular = -dy
                        dy_perpendicular = dx

                        # Desplazar la posición del texto
                        distancia_desplazamiento = 25 # Ajusta este valor para controlar la distancia de desplazamiento
                        x_desplazado = int((x1 + x2) / 2 + dx_perpendicular * distancia_desplazamiento)
                        y_desplazado = int((y1 + y2) / 2 + dy_perpendicular * distancia_desplazamiento)

                        textos.append((x_desplazado, y_desplazado, texto_distancia, (0, 0, 255), cv2.FONT_HERSHEY_SIMPLEX, 1.5))

                        puntos = []
                        actualizar_imagen()

                elif event == cv2.EVENT_RBUTTONDOWN:
                    arrastrando = True
                    x_anterior, y_anterior = x, y

                elif event == cv2.EVENT_MOUSEMOVE:
                    if arrastrando:
                        delta_x = x - x_anterior
                        delta_y = y - y_anterior
                        x_offset += delta_x
                        y_offset += delta_y
                        x_anterior, y_anterior = x, y
                        actualizar_imagen()

                    if len(puntos) == 1:
                        x_anterior, y_anterior = x, y
                        actualizar_imagen()

                elif event == cv2.EVENT_RBUTTONUP:
                    arrastrando = False

                elif event == cv2.EVENT_MOUSEWHEEL:
                    if flags > 0:
                        nuevo_zoom = zoom_factor * 1.1
                    else:
                        nuevo_zoom = zoom_factor / 1.1
                    nuevo_zoom = max(1.0, nuevo_zoom)

                    x_raton, y_raton = x, y
                    x_offset = x_raton - (x_raton - x_offset) * (nuevo_zoom / zoom_factor)
                    y_offset = y_raton - (y_raton - y_offset) * (nuevo_zoom / zoom_factor)
                    zoom_factor = nuevo_zoom

                    actualizar_imagen()

            cv2.setMouseCallback("Medir Distancia", clic_derecho)
            cv2.imshow("Medir Distancia", imagen_original)
            cv2.waitKey(0)

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Ocurrió un error: {e}")

    def conteo_manual(self):
        try:
            seleccion = self.lista_imagenes.currentRow()
            if seleccion < 0:
                QMessageBox.critical(self, "Error", "Por favor, selecciona una imagen de la lista.")
                return

            ruta = self.rutas_imagenes[seleccion]
            imagen = cv2.imread(ruta, cv2.IMREAD_COLOR)
            if imagen is None:
                QMessageBox.critical(self, "Error", "No se pudo cargar la imagen.")
                return

            cv2.namedWindow("Conteo Manual", cv2.WINDOW_NORMAL)
            puntos = []
            contador = 1
            imagen_original = imagen.copy()
            zoom_factor = 1.0
            x_offset, y_offset = 0, 0
            arrastrando = False
            x_anterior, y_anterior = 0, 0

            def transformar_coordenadas(x, y):
                return int(x * zoom_factor + x_offset), int(y * zoom_factor + y_offset)

            def invertir_transformacion(x, y):
                return int((x - x_offset) / zoom_factor), int((y - y_offset) / zoom_factor)

            def dibujar_cruz(imagen, x, y, color, grosor, texto):
                longitud = 10
                cv2.line(imagen, (x - longitud, y), (x + longitud, y), color, grosor)
                cv2.line(imagen, (x, y - longitud), (x, y + longitud), color, grosor)
                cv2.putText(imagen, str(texto), (x + longitud + 5, y + longitud + 5), cv2.FONT_HERSHEY_DUPLEX, 1.5, color, 2)

            def actualizar_imagen():
                nonlocal imagen_original, zoom_factor, x_offset, y_offset, puntos

                M = np.float32([
                    [zoom_factor, 0, x_offset],
                    [0, zoom_factor, y_offset]
                ])
                alto, ancho = imagen_original.shape[:2]
                imagen_transformada = cv2.warpAffine(imagen_original, M, (ancho, alto))

                for i, (x, y) in enumerate(puntos, start=1):
                    x_t, y_t = transformar_coordenadas(x, y)
                    dibujar_cruz(imagen_transformada, x_t, y_t, (0, 0, 255), 2, i)

                cv2.imshow("Conteo Manual", imagen_transformada)

            def clic_derecho(event, x, y, flags, param):
                nonlocal puntos, contador, arrastrando, x_anterior, y_anterior, x_offset, y_offset, zoom_factor

                x_original, y_original = invertir_transformacion(x, y)

                if event == cv2.EVENT_LBUTTONDOWN:
                    puntos.append((x_original, y_original))
                    contador += 1
                    actualizar_imagen()

                elif event == cv2.EVENT_RBUTTONDOWN:
                    arrastrando = True
                    x_anterior, y_anterior = x, y

                elif event == cv2.EVENT_MOUSEMOVE:
                    if arrastrando:
                        delta_x = x - x_anterior
                        delta_y = y - y_anterior
                        x_offset += delta_x
                        y_offset += delta_y
                        x_anterior, y_anterior = x, y
                        actualizar_imagen()

                elif event == cv2.EVENT_RBUTTONUP:
                    arrastrando = False

                elif event == cv2.EVENT_MOUSEWHEEL:
                    if flags > 0:
                        nuevo_zoom = zoom_factor * 1.1
                    else:
                        nuevo_zoom = zoom_factor / 1.1
                    nuevo_zoom = max(1.0, nuevo_zoom)

                    x_raton, y_raton = x, y
                    x_offset = x_raton - (x_raton - x_offset) * (nuevo_zoom / zoom_factor)
                    y_offset = y_raton - (y_raton - y_offset) * (nuevo_zoom / zoom_factor)
                    zoom_factor = nuevo_zoom

                    actualizar_imagen()

            cv2.setMouseCallback("Conteo Manual", clic_derecho)
            cv2.imshow("Conteo Manual", imagen_original)
            cv2.waitKey(0)

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Ocurrió un error: {e}")

    def medir_area_intensidad(self):
        try:
            seleccion = self.lista_imagenes.currentRow()
            if seleccion < 0:
                QMessageBox.critical(self, "Error", "Por favor, selecciona una imagen de la lista.")
                return

            ruta = self.rutas_imagenes[seleccion]
            imagen = cv2.imread(ruta, cv2.IMREAD_COLOR)
            if imagen is None:
                QMessageBox.critical(self, "Error", "No se pudo cargar la imagen.")
                return

            # Pedir la escala en µm/px
            um_per_px, ok = QInputDialog.getDouble(self, "Escala", "Ingresa la escala en µm/px:", decimals=2)
            if not ok:
                return

            um_per_px_str = str(um_per_px).replace(",", ".") #Reemplazar comas por puntos.
            try:
                um_per_px = float(um_per_px_str)
            except ValueError:
                QMessageBox.critical(self, "Error", "Escala inválida. Por favor, ingresa un número válido.")
                return

            if um_per_px <= 0:
                QMessageBox.critical(self, "Error", "La escala debe ser mayor que cero.")
                return

            self.um_per_px = um_per_px

            cv2.namedWindow("Medir Área e Intensidad", cv2.WINDOW_NORMAL)
            figuras = []
            imagen_original = imagen.copy()
            zoom_factor = 1.0
            x_offset, y_offset = 0, 0
            arrastrando = False
            x_inicio, y_inicio = 0, 0
            arrastrando_desplazamiento = False
            figura_seleccionada = None
            rotando = False

            def calcular_area_intensidad(figura):
                if figura is None:
                    return 0, 0

                mascara = np.zeros_like(imagen[:, :, 0], dtype=np.uint8)
                x1, y1 = figura["puntos"][0]
                x2, y2 = figura["puntos"][1]
                cv2.rectangle(mascara, (x1, y1), (x2, y2), 255, -1)

                area_px = cv2.countNonZero(mascara)
                area_um2 = area_px * (um_per_px ** 2)

                imagen_gris = cv2.cvtColor(imagen, cv2.COLOR_BGR2GRAY)
                intensidad_media = cv2.mean(imagen_gris, mask=mascara)[0]

                return area_um2, intensidad_media

            def actualizar_imagen():
                nonlocal imagen_original, figuras, zoom_factor, x_offset, y_offset

                M = np.float32([
                    [zoom_factor, 0, x_offset],
                    [0, zoom_factor, y_offset]
                ])
                alto, ancho = imagen_original.shape[:2]
                imagen_transformada = cv2.warpAffine(imagen_original, M, (ancho, alto))

                for figura_actual in figuras:
                    x1, y1 = figura_actual["puntos"][0]
                    x2, y2 = figura_actual["puntos"][1]

                    x1_t, y1_t = transformar_coordenadas(x1, y1)
                    x2_t, y2_t = transformar_coordenadas(x2, y2)

                    color = (0, 0, 255) if figura_actual == figura_seleccionada else (0, 255, 0)
                    cv2.rectangle(imagen_transformada, (x1_t, y1_t), (x2_t, y2_t), color, 2)

                    # Marcas de redimensionamiento (líneas más gruesas)
                    if figura_actual == figura_seleccionada:
                        centro_x_arriba, centro_y_arriba = (x1_t + x2_t) // 2, y1_t
                        centro_x_abajo, centro_y_abajo = (x1_t + x2_t) // 2, y2_t
                        centro_x_izquierda, centro_y_izquierda = x1_t, (y1_t + y2_t) // 2
                        centro_x_derecha, centro_y_derecha = x2_t, (y1_t + y2_t) // 2

                        cv2.line(imagen_transformada, (centro_x_arriba - 10, centro_y_arriba), (centro_x_arriba + 10, centro_y_arriba), color, 10)  # Aumentado a 5
                        cv2.line(imagen_transformada, (centro_x_abajo - 10, centro_y_abajo), (centro_x_abajo + 10, centro_y_abajo), color, 10)  # Aumentado a 5
                        cv2.line(imagen_transformada, (centro_x_izquierda, centro_y_izquierda - 10), (centro_x_izquierda, centro_y_izquierda + 10), color, 10)  # Aumentado a 5
                        cv2.line(imagen_transformada, (centro_x_derecha, centro_y_derecha - 10), (centro_x_derecha, centro_y_derecha + 10), color, 10)  # Aumentado a 5

                    area_um2, intensidad_media = calcular_area_intensidad(figura_actual)
                    texto = f"Area: {area_um2:.2f} um2, Intensidad: {intensidad_media:.2f}"
                    posicion_texto = (x1_t, y1_t - 10)
                    cv2.putText(imagen_transformada, texto, posicion_texto, cv2.FONT_HERSHEY_SIMPLEX, 0.8, color, 2)

                cv2.imshow("Medir Área e Intensidad", imagen_transformada)

            def transformar_coordenadas(x, y):
                return int(x * zoom_factor + x_offset), int(y * zoom_factor + y_offset)

            def invertir_transformacion(x, y):
                return int((x - x_offset) / zoom_factor), int((y - y_offset) / zoom_factor)

            def clic_derecho(event, x, y, flags, param):
                nonlocal figuras, arrastrando, x_inicio, y_inicio, zoom_factor, x_offset, y_offset, arrastrando_desplazamiento, figura_seleccionada, rotando

                x_original, y_original = invertir_transformacion(x, y)

                if event == cv2.EVENT_LBUTTONDOWN:
                    figura_seleccionada = None
                    clic_en_rectangulo = False
                    for figura in figuras:
                        x1, y1 = figura["puntos"][0]
                        x2, y2 = figura["puntos"][1]
                        if x1 <= x_original <= x2 and y1 <= y_original <= y2:
                            figura_seleccionada = figura
                            clic_en_rectangulo = True
                            break

                    if figura_seleccionada:
                        arrastrando = True
                        x_inicio, y_inicio = x_original, y_original
                        # Verificar si el clic está cerca de las marcas para redimensionar
                        centro_x_arriba, centro_y_arriba = (figura_seleccionada["puntos"][0][0] + figura_seleccionada["puntos"][1][0]) // 2, figura_seleccionada["puntos"][0][1]
                        centro_x_abajo, centro_y_abajo = (figura_seleccionada["puntos"][0][0] + figura_seleccionada["puntos"][1][0]) // 2, figura_seleccionada["puntos"][1][1]
                        centro_x_izquierda, centro_y_izquierda = figura_seleccionada["puntos"][0][0], (figura_seleccionada["puntos"][0][1] + figura_seleccionada["puntos"][1][1]) // 2
                        centro_x_derecha, centro_y_derecha = figura_seleccionada["puntos"][1][0], (figura_seleccionada["puntos"][0][1] + figura_seleccionada["puntos"][1][1]) // 2

                        if abs(x_original - centro_x_arriba) < 15 and abs(y_original - centro_y_arriba) < 5:
                            figura_seleccionada["redimensionar"] = "arriba"
                        elif abs(x_original - centro_x_abajo) < 15 and abs(y_original - centro_y_abajo) < 5:
                            figura_seleccionada["redimensionar"] = "abajo"
                        elif abs(x_original - centro_x_izquierda) < 5 and abs(y_original - centro_y_izquierda) < 15:
                            figura_seleccionada["redimensionar"] = "izquierda"
                        elif abs(x_original - centro_x_derecha) < 5 and abs(y_original - centro_y_derecha) < 15:
                            figura_seleccionada["redimensionar"] = "derecha"
                        else:
                            figura_seleccionada["redimensionar"] = None
                    else:
                        figura_actual = {"tipo": "rectangulo", "puntos": [(x_original, y_original), (x_original, y_original)]}
                        figuras.append(figura_actual)
                        arrastrando = True
                        x_inicio, y_inicio = x_original, y_original

                elif event == cv2.EVENT_MOUSEMOVE:
                    if arrastrando and figura_seleccionada:
                        delta_x = x_original - x_inicio
                        delta_y = y_original - y_inicio
                        if figura_seleccionada["redimensionar"] == "izquierda":
                                figura_seleccionada["puntos"][0] = (figura_seleccionada["puntos"][0][0] + delta_x, figura_seleccionada["puntos"][0][1])
                        elif figura_seleccionada["redimensionar"] == "derecha":
                                figura_seleccionada["puntos"][1] = (figura_seleccionada["puntos"][1][0] + delta_x, figura_seleccionada["puntos"][1][1])
                        elif figura_seleccionada["redimensionar"] == "arriba":
                                figura_seleccionada["puntos"][0] = (figura_seleccionada["puntos"][0][0], figura_seleccionada["puntos"][0][1] + delta_y)
                        elif figura_seleccionada["redimensionar"] == "abajo":
                                figura_seleccionada["puntos"][1] = (figura_seleccionada["puntos"][1][0], figura_seleccionada["puntos"][1][1] + delta_y)
                        else:
                            figura_seleccionada["puntos"][0] = (figura_seleccionada["puntos"][0][0] + delta_x, figura_seleccionada["puntos"][0][1] + delta_y)
                            figura_seleccionada["puntos"][1] = (figura_seleccionada["puntos"][1][0] + delta_x, figura_seleccionada["puntos"][1][1] + delta_y)
                        x_inicio, y_inicio = x_original, y_original
                        actualizar_imagen()
                        
                    elif arrastrando:
                        figura_actual = figuras[-1]
                        figura_actual["puntos"][1] = (x_original, y_original)
                        actualizar_imagen()
                    elif arrastrando_desplazamiento:
                        delta_x = x - x_inicio
                        delta_y = y - y_inicio
                        x_offset += delta_x
                        y_offset += delta_y
                        x_inicio, y_inicio = x, y
                        actualizar_imagen()

                elif event == cv2.EVENT_LBUTTONUP:
                    arrastrando = False
                    if figura_seleccionada:
                        figura_seleccionada["redimensionar"] = None
                    actualizar_imagen()

                elif event == cv2.EVENT_RBUTTONDOWN:
                    arrastrando_desplazamiento = True
                    x_inicio, y_inicio = x, y

                elif event == cv2.EVENT_RBUTTONUP:
                    arrastrando_desplazamiento = False

                elif event == cv2.EVENT_MOUSEWHEEL:
                    zoom_anterior = zoom_factor
                    if flags > 0:
                        zoom_factor *= 1.1
                    else:
                        zoom_factor /= 1.1
                    zoom_factor = max(1.0, zoom_factor)
                    delta_zoom = zoom_factor - zoom_anterior
                    x_offset -= int(delta_zoom * (x - x_offset))
                    y_offset -= int(delta_zoom * (y - y_offset))
                    actualizar_imagen()

            cv2.setMouseCallback("Medir Área e Intensidad", clic_derecho)
            cv2.imshow("Medir Área e Intensidad", imagen_original)
            cv2.waitKey(0)

        except Exception as e:
            QMessageBox.critical(self, "Error")

    def procesar_todas_imagenes(self):
        if not self.rutas_imagenes:
            QMessageBox.warning(self, "Advertencia", "No hay imágenes cargadas.")
            return

        # ===============================
        # ESCALA Y RANGO
        # ===============================
        if hasattr(self, 'escala_predeterminada') and hasattr(self, 'rango_longitud_predeterminado'):
            usar_predeterminados = QMessageBox.question(
                self,
                "Usar Parámetros Predeterminados",
                "¿Deseas usar los parámetros predeterminados?\n"
                f"Escala: {self.escala_predeterminada} µm/px\n"
                f"Rango: {self.rango_longitud_predeterminado[0]} - {self.rango_longitud_predeterminado[1]} µm",
                QMessageBox.Yes | QMessageBox.No,
            )

            if usar_predeterminados == QMessageBox.Yes:
                self.um_per_px = self.escala_predeterminada
                self.rango_longitud = self.rango_longitud_predeterminado
            else:
                self.pedir_escala_y_rango()
        else:
            self.pedir_escala_y_rango()

        # ===============================
        # PREPROCESADO
        # ===============================
        opcion, ok = QInputDialog.getInt(
            self,
            "Tipo de Preprocesamiento",
            "1. Blanco sobre negro\n2. Negro sobre blanco",
            min=1, max=2
        )

        if not ok:
            return

        # ===============================
        # TXT OBLIGATORIO
        # ===============================
        ruta_txt, _ = QFileDialog.getOpenFileName(
            self,
            "Selecciona el archivo espectral (.txt)",
            "",
            "TXT (*.txt)"
        )

        if not ruta_txt:
            QMessageBox.warning(self, "Cancelado", "No se seleccionó archivo espectral.")
            return

        # ===============================
        # CARGA TXT UNA VEZ
        # ===============================
        espectro_txt = np.loadtxt(ruta_txt)
        espectro_txt = np.nan_to_num(espectro_txt)

        # quitar eje Y
        espectro_txt = espectro_txt[:, 1:]

        carpeta_guardado = QFileDialog.getExistingDirectory(
            self,
            "Selecciona carpeta de guardado"
        )

        if not carpeta_guardado:
            return

        self.progress.setValue(0)

        resultados_particulas = []
        resultados_fibras = []

        # ===============================
        # LOOP IMÁGENES
        # ===============================
        for idx, ruta in enumerate(self.rutas_imagenes):

            try:
                self.progress.setValue(int((idx / len(self.rutas_imagenes)) * 100))
                QApplication.processEvents()

                imagen = cv2.imread(ruta, cv2.IMREAD_GRAYSCALE)
                if imagen is None:
                    continue

                # ===============================
                # CHECK SIZE TXT vs BMP
                # ===============================
                if espectro_txt.shape[0] != imagen.shape[0]:
                    QMessageBox.warning(
                        self,
                        "Aviso de tamaño",
                        f"Diferencia detectada:\nBMP: {imagen.shape}\nTXT: {espectro_txt.shape}\nSe ajustará automáticamente"
                    )

                # ===============================
                # AJUSTE TXT A IMAGEN
                # ===============================
                espectro_ajustado = cv2.resize(
                    espectro_txt,
                    (imagen.shape[1], imagen.shape[0]),
                    interpolation=cv2.INTER_LINEAR
                )

                # ===============================
                # PREPROCESADO
                # ===============================
                if opcion == 1:
                    _, umbralizada = cv2.threshold(
                        imagen, 127, 255, cv2.THRESH_BINARY + cv2.THRESH_TRIANGLE
                    )
                else:
                    _, umbralizada = cv2.threshold(
                        imagen, 45, 255, cv2.THRESH_BINARY_INV
                    )

                kernel = np.ones((3, 3), np.uint8)
                umbralizada = cv2.morphologyEx(umbralizada, cv2.MORPH_CLOSE, kernel)

                contornos, _ = cv2.findContours(
                    umbralizada, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE
                )

                imagen_color = cv2.cvtColor(imagen, cv2.COLOR_GRAY2BGR)

                rango_fibras = (350, 1500)

                datos_particulas = []
                datos_fibras = []

                # ===============================
                # CONTORNOS
                # ===============================
                for i, contorno in enumerate(contornos, start=1):

                    area_px = cv2.contourArea(contorno)
                    perimetro_px = cv2.arcLength(contorno, True)

                    if area_px <= 1:
                        continue

                    area_um2 = area_px * (self.um_per_px ** 2)
                    perimetro_um = perimetro_px * self.um_per_px

                    (_, _), radius = cv2.minEnclosingCircle(contorno)
                    diametro_max_um = 2 * radius * self.um_per_px

                    rect = cv2.minAreaRect(contorno)
                    ancho, alto = rect[1]
                    diametro_min_um = min(ancho, alto) * self.um_per_px

                    x, y, w, h = cv2.boundingRect(contorno)

                    relacion_aspecto = (
                        diametro_min_um / diametro_max_um if diametro_max_um else 0
                    )

                    circularidad = (
                        (4 * np.pi * area_px) / (perimetro_px ** 2)
                        if perimetro_px else 0
                    )

                    # ===============================
                    # ROI → ESPECTRO REAL
                    # ===============================
                    mascara = np.zeros(imagen.shape, dtype=np.uint8)
                    cv2.drawContours(mascara, [contorno], -1, 255, -1)

                    valores = espectro_ajustado[mascara == 255]

                    datos_espectrales = None
                    if len(valores) > 0:
                        datos_espectrales = {
                            "media": float(np.mean(valores)),
                            "max": float(np.max(valores)),
                            "min": float(np.min(valores)),
                            "std": float(np.std(valores))
                        }

                    # ===============================
                    # FIBRAS
                    # ===============================
                    if relacion_aspecto < 0.3 or circularidad < 0.12:

                        if rango_fibras[0] <= diametro_max_um <= rango_fibras[1]:

                            dato = (
                                os.path.basename(ruta),
                                i,
                                diametro_max_um,
                                diametro_min_um,
                                area_um2,
                                perimetro_um,
                                relacion_aspecto,
                                circularidad
                            )

                            if datos_espectrales:
                                dato += (
                                    datos_espectrales["media"],
                                    datos_espectrales["max"],
                                    datos_espectrales["min"],
                                    datos_espectrales["std"]
                                )

                            datos_fibras.append(dato)

                            cv2.drawContours(imagen_color, [contorno], -1, (0, 0, 255), 2)

                    # ===============================
                    # PARTÍCULAS
                    # ===============================
                    else:
                        if self.rango_longitud[0] <= diametro_max_um <= self.rango_longitud[1]:

                            dato = (
                                os.path.basename(ruta),
                                i,
                                diametro_max_um,
                                diametro_min_um,
                                area_um2,
                                perimetro_um,
                                relacion_aspecto,
                                circularidad
                            )

                            if datos_espectrales:
                                dato += (
                                    datos_espectrales["media"],
                                    datos_espectrales["max"],
                                    datos_espectrales["min"],
                                    datos_espectrales["std"]
                                )

                            datos_particulas.append(dato)

                            cv2.drawContours(imagen_color, [contorno], -1, (0, 255, 0), 2)

                # ===============================
                # GUARDADO
                # ===============================
                nombre_archivo = os.path.basename(ruta)
                ruta_guardado = os.path.join(carpeta_guardado, f"procesada_{nombre_archivo}")
                cv2.imwrite(ruta_guardado, imagen_color)

                resultados_particulas.extend(datos_particulas)
                resultados_fibras.extend(datos_fibras)

            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error en {ruta}: {e}")

        # ===============================
        # RESULTADOS
        # ===============================
        self.mostrar_tabla_consolidada(resultados_particulas, resultados_fibras)
        self.mostrar_resumen_consolidado(resultados_particulas, resultados_fibras)
        self.progress.setValue(100)
        QApplication.processEvents()


    def mostrar_imagen_procesada(self, imagen_color, nombre_archivo):
        # Variables para el zoom, desplazamiento y estado de arrastre
        zoom_factor = 1.0
        x_offset, y_offset = 0, 0
        arrastrando = False
        x_inicio, y_inicio = 0, 0

        def actualizar_imagen():
            nonlocal imagen_color, zoom_factor, x_offset, y_offset

            # Crear una matriz de transformación para el zoom y el desplazamiento
            M = np.float32([
                [zoom_factor, 0, x_offset],
                [0, zoom_factor, y_offset]
            ])

            # Aplicar la transformación a la imagen original
            alto, ancho = imagen_color.shape[:2]
            imagen_transformada = cv2.warpAffine(imagen_color, M, (ancho, alto))

            # Redimensionar la imagen para que se ajuste a la ventana
            alto_ventana, ancho_ventana = 800, 1200  # Tamaño de la ventana
            imagen_redimensionada = cv2.resize(imagen_transformada, (ancho_ventana, alto_ventana))

            # Mostrar la imagen actualizada
            cv2.imshow(f"Imagen Procesada - {nombre_archivo}", imagen_redimensionada)

        def clic_derecho(event, x, y, flags, param):
            nonlocal zoom_factor, x_offset, y_offset, arrastrando, x_inicio, y_inicio

            # Ajustar las coordenadas del mouse al tamaño de la imagen original
            x_original = int(x * (imagen_color.shape[1] / 1200))  # 1200 es el ancho de la ventana
            y_original = int(y * (imagen_color.shape[0] / 800))   # 800 es el alto de la ventana

            if event == cv2.EVENT_MOUSEWHEEL:
                # Hacer zoom con la rueda del ratón
                if flags > 0:  # Rueda hacia arriba (zoom in)
                    nuevo_zoom = zoom_factor * 1.1
                else:  # Rueda hacia abajo (zoom out)
                    nuevo_zoom = zoom_factor / 1.1
                nuevo_zoom = max(1.0, nuevo_zoom)  # Evitar zoom menor a 1

                # Ajustar los offsets para que el zoom se centre en el cursor
                x_raton, y_raton = x_original, y_original
                x_offset = x_raton - (x_raton - x_offset) * (nuevo_zoom / zoom_factor)
                y_offset = y_raton - (y_raton - y_offset) * (nuevo_zoom / zoom_factor)
                zoom_factor = nuevo_zoom

                actualizar_imagen()

            elif event == cv2.EVENT_LBUTTONDOWN:
                # Iniciar arrastre
                arrastrando = True
                x_inicio, y_inicio = x_original, y_original

            elif event == cv2.EVENT_MOUSEMOVE:
                if arrastrando:
                    # Calcular el desplazamiento
                    delta_x = x_original - x_inicio
                    delta_y = y_original - y_inicio
                    x_offset += delta_x
                    y_offset += delta_y
                    x_inicio, y_inicio = x_original, y_original
                    actualizar_imagen()

            elif event == cv2.EVENT_LBUTTONUP:
                # Detener arrastre
                arrastrando = False

        cv2.namedWindow(f"Imagen Procesada - {nombre_archivo}", cv2.WINDOW_NORMAL)
        cv2.resizeWindow(f"Imagen Procesada - {nombre_archivo}", 1200, 800)  # Tamaño de la ventana
        cv2.setMouseCallback(f"Imagen Procesada - {nombre_archivo}", clic_derecho)  # Enlazar evento del ratón
        actualizar_imagen()  # Mostrar la imagen inicial
        cv2.waitKey(1)  # Esperar hasta que se cierre la ventana

    def mostrar_tabla_consolidada(self, resultados_particulas, resultados_fibras):
        ventana_tabla = QDialog(self)
        ventana_tabla.setWindowTitle("Resultados Consolidados")
        ventana_tabla.resize(1200, 800)
        layout = QVBoxLayout(ventana_tabla)

        notebook = QTabWidget()
        layout.addWidget(notebook)

        def guardar_excel(datos_particulas, datos_fibras, nombre_archivo):
            try:
                libro_excel = openpyxl.Workbook()

                # Hoja para partículas
                hoja_particulas = libro_excel.active
                hoja_particulas.title = "Partículas"
                encabezado_particulas = ["Imagen", "ID", "Diámetro (µm)", "Diámetro Mín (µm)", "Área (µm²)", "Perímetro (µm)", "Relación de Aspecto", "Circularidad"]
                hoja_particulas.append(encabezado_particulas)
                for fila in datos_particulas:
                    hoja_particulas.append(fila[:8])  # Guardar solo hasta Circularidad

                # Hoja para fibras
                hoja_fibras = libro_excel.create_sheet("Fibras")
                encabezado_fibras = ["Imagen", "ID", "Diámetro (µm)", "Diámetro Mín (µm)", "Área (µm²)", "Perímetro (µm)", "Relación de Aspecto", "Circularidad"]
                hoja_fibras.append(encabezado_fibras)
                for fila in datos_fibras:
                    hoja_fibras.append(fila[:8])  # Guardar solo hasta Circularidad

                libro_excel.save(nombre_archivo)
                QMessageBox.information(ventana_tabla, "Éxito", f"Datos guardados en {nombre_archivo}")
            except Exception as e:
                QMessageBox.critical(ventana_tabla, "Error", f"No se pudieron guardar los datos: {e}")

        # Pestaña de partículas
        frame_particulas = QWidget()
        layout_particulas = QVBoxLayout(frame_particulas)
        tabla_particulas = QTableWidget()
        tabla_particulas.setColumnCount(8)
        tabla_particulas.setHorizontalHeaderLabels(["Imagen", "ID", "Diámetro (µm)", "Diámetro Mín (µm)", "Área (µm²)", "Perímetro (µm)", "Relación de Aspecto", "Circularidad"])

        tabla_particulas.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        tabla_particulas.verticalHeader().setDefaultSectionSize(30)

        for dato in resultados_particulas:
            tabla_particulas.insertRow(tabla_particulas.rowCount())
            for i, valor in enumerate(dato[:8]):  # Mostrar solo hasta Circularidad
                tabla_particulas.setItem(tabla_particulas.rowCount() - 1, i, QTableWidgetItem(str(valor)))
        layout_particulas.addWidget(tabla_particulas)

        btn_guardar_particulas = QPushButton("Guardar Excel")
        btn_guardar_particulas.clicked.connect(lambda: guardar_excel(resultados_particulas, resultados_fibras, QFileDialog.getSaveFileName(ventana_tabla, "Guardar Datos en Excel", "", "Archivos Excel (*.xlsx)")[0]))
        layout_particulas.addWidget(btn_guardar_particulas)

        notebook.addTab(frame_particulas, "Partículas")

        # Pestaña de fibras
        frame_fibras = QWidget()
        layout_fibras = QVBoxLayout(frame_fibras)
        tabla_fibras = QTableWidget()
        tabla_fibras.setColumnCount(8)
        tabla_fibras.setHorizontalHeaderLabels(["Imagen", "ID", "Diámetro (µm)", "Diámetro Mín (µm)", "Área (µm²)", "Perímetro (µm)", "Relación de Aspecto", "Circularidad"])

        tabla_fibras.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        tabla_fibras.verticalHeader().setDefaultSectionSize(30)

        for dato in resultados_fibras:
            tabla_fibras.insertRow(tabla_fibras.rowCount())
            for i, valor in enumerate(dato[:8]):  # Mostrar solo hasta Circularidad
                tabla_fibras.setItem(tabla_fibras.rowCount() - 1, i, QTableWidgetItem(str(valor)))
        layout_fibras.addWidget(tabla_fibras)

        notebook.addTab(frame_fibras, "Fibras")

        ventana_tabla.exec_()

    def mostrar_resumen_consolidado(self, resultados_particulas, resultados_fibras):
        ventana_resumen = QDialog(self)
        ventana_resumen.setWindowTitle("Resumen Consolidado")
        layout = QVBoxLayout(ventana_resumen)

        lbl_particulas = QLabel(f"Total de Partículas Detectadas: {len(resultados_particulas)}")
        lbl_fibras = QLabel(f"Total de Fibras Detectadas: {len(resultados_fibras)}")
        layout.addWidget(lbl_particulas)
        layout.addWidget(lbl_fibras)

        ventana_resumen.exec_()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    try:
        window = ParticleCounterApp()
        window.show()
        sys.exit(app.exec_())
    except Exception as e:
        logging.critical(f"Error global: {e}")
        QMessageBox.critical(None, "Error Crítico", f"Ocurrió un error crítico: {e}")
